import time


from selenium import webdriver
import selenium
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.action_chains import ActionChains
from bs4 import BeautifulSoup
from furl import furl
from typing import List
from openpyxl import Workbook
from openpyxl import load_workbook


class Automation:
    def __init__(self):
        self.chrome_options = Options()
        self.chrome_options.add_argument("--window-size=1920,1080")
        self.chrome_options.add_argument("--disable-extensions")
        self.chrome_options.add_argument("--proxy-server='direct://'")
        self.chrome_options.add_argument("--proxy-bypass-list=*")
        self.chrome_options.add_argument("--start-maximized")
        # self.chrome_options.add_argument("--headless")
        self.chrome_options.add_argument("--disable-gpu")
        self.chrome_options.add_argument("--disable-dev-shm-usage")
        self.chrome_options.add_argument("--no-sandbox")
        self.chrome_options.add_argument("--ignore-certificate-errors")
        self.amazon_appstore_link = 'https://sellercentral.amazon.com/selling-partner-appstore'
        self.categories_to_scrape = [
            "Product Research and Scouting",
            "Listing",
            "Automated Pricing",
            "Inventory and Order Management",
            "Shipping Solutions",
            "Advertising Optimization",
            "Promotions",
            "Feedback and Reviews",
            "Buyer-Seller Messaging Service",
            "Analytics and Reporting",
            "Accounting",
            "Funding and Credit",
            "Taxes",
            "Disbursement Solutions",
            "Ecommerce Solution Connectors",
            "Full Service Solutions"
        ]
        self.driver = webdriver.Chrome(
            service=ChromeService(ChromeDriverManager().install()),
            options=self.chrome_options,
        )

    def scrape_apps_from_category(self, category_element: selenium.webdriver.remote.webelement.WebElement):
        category_name = category_element.text
        category_element.click()

        time.sleep(2)
        apps = self.scrape_apps_from_page(category_name)

        while self.go_on_next_page():
            new_batch = self.scrape_apps_from_page(category_name)
            print(new_batch)
            apps += new_batch
        return apps

    def go_on_next_page(self) -> bool:
        link = self.driver.current_url
        f = furl(link)
        number = f.args['pageNumber']
        f.args['pageNumber'] = int(number) + 1
        self.driver.get(f.url)
        time.sleep(4)
        try:
            if self.driver.find_element(By.ID, 'mas-apps-list-no-results-header'):
                return False
        except NoSuchElementException:
            try:
                if self.driver.find_element(By.ID, 'mas-apps-list-tile-grid'):
                    return True
            except NoSuchElementException:
                return False
        return False

    def scrape_apps_from_page(self, category_name: str) -> List[dict]:
        time.sleep(2)

        html = self.driver.page_source
        print("I am called!")
        soup = BeautifulSoup(html, features="html.parser")
        print("soup: ", soup)
        apps = soup.find_all('div', {'class': 'solution-tile-content-container'})

        apps_list = []

        for app in apps:
            link_to_app = app.find('a', href=True)['href']
            app_name = app.find('div', {'class': 'solution-tile-name'}).text
            short_description = app.find('span', {'class': 'solution-tile-description-text'}).text
            apps_list.append(
                {'link_to_app': f"https://sellercentral.amazon.com{link_to_app}", 'app_name': app_name,
                 'short_description': short_description, 'category_name': category_name}
            )

        return apps_list

    def run(self):
        """
        CATEGORIES:
        Product Research and Scouting
        Listing
        Automated Pricing
        Inventory and Order Management
        Shipping Solutions
        Advertising Optimization
        Promotions
        Feedback and Reviews
        Buyer-Seller Messaging Service
        Analytics and Reporting
        Accounting
        Funding and Credit
        Taxes
        Disbursement Solutions
        Ecommerce Solution Connectors
        Full Service Solutions
        """

        self.driver.get(self.amazon_appstore_link)
        time.sleep(5)

        filename = "amazon_apps.xlsx"

        for category in self.categories_to_scrape:
            elements = self.driver.find_elements(By.CLASS_NAME, "category-list-item")
            print(elements)
            print('----')
            print(category)

            current_categories = [x for x in elements if x.text != '']
            # print(current_categories)
            chosen_category = [x for x in current_categories if x.text == category][0]
            new_rows = self.scrape_apps_from_category(chosen_category)

            try:
                wb = load_workbook(filename)
                ws = wb.worksheets[0]  # select first worksheet
            except FileNotFoundError:
                headers_row = ['app_name', 'category_name', 'link_to_app', 'short_description']
                wb = Workbook()
                ws = wb.active
                ws.append(headers_row)

            field_names = ['app_name', 'category_name', 'link_to_app', 'short_description']

            for row in new_rows:
                values = (row[k] for k in field_names)
                ws.append(values)
                wb.save(filename)
            self.driver.get(self.amazon_appstore_link)
            time.sleep(5)


if __name__ == '__main__':
    print("Start")
    x = Automation()
    x.run()
    print("End")